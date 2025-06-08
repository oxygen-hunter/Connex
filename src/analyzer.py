# -*- coding:utf-8 -*-

import json, os
from easydict import EasyDict as ed
import openpyxl
from openpyxl.styles import Alignment
import typing as T

from src import config as C , tools as TL, doc_info as DOC
from src.gt_generator import _generate_tx_hash_to_key_info_dict,_generate_token_info_from_key_info
from src.rules import get_rules
from collections import defaultdict as dd

def _write_paired_token_to_excel(paired_token:T.List[T.Dict], path:str):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    r = 0
    c = 0
    title = ['id', 'chain', 'token_addr', 'token_name', 'token_symbol', 'count', 'fee_rate', 'fee_amount']
    for c in range(0, len(title)):
        ws.cell(r+1, c+1).value = title[c]
    r += 1
    for one_pair in paired_token:
        token1 = one_pair['token1']
        token2 = one_pair['token2']
        ws.cell(r+1, 1).value = (r+1)//2
        for c in range(1, len(token1.keys())+1):
            ws.cell(r+1, c+1).value = token1[title[c]]
        ws.cell(r+1, c+1+1).value = one_pair['count']
        ws.cell(r+1, c+1+1+1).value = str(one_pair['fee_rate'])
        ws.cell(r+1, c+1+1+1+1).value = str(one_pair['fee_amount'])
        r += 1
        for c in range(1, len(token2.keys())+1):
            try:
                ws.cell(r+1, c+1).value = token2[title[c]]
            except openpyxl.utils.exceptions.IllegalCharacterError:
                ws.cell(r+1, c+1).value = "!!illegal"
        r += 1
    
    # auto adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = (length+2) * 1.2
    
    # left align
    for column_cells in ws.columns:
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='left')

    wb.save(path)
        

treshold_related_case = set()

def classify_violation_rules(
        vl:str,
        _GT, 
        violation_rules_cnt:dict, 
        violation_rules_statis:dict,
        violation_rules_class:dict,
        real_ts1=None, real_ts2=None):

    violation_rules_cnt[vl] += 1
    if vl == 'rule2_timestamp' and 'generate_key_info_fail' not in _GT['violate_rule']:
        tw = _GT['pair_info']['time_window']
        violation_rules_statis['timewindow'].append(tw)
        real_tw = real_ts2 - real_ts1
        if real_tw != tw: 
            violation_rules_class['timewindow']['extract_wrong'] += 1
        elif tw <= 0 : 
            violation_rules_class['timewindow']['0'] += 1
        elif tw > 1700000000:
            violation_rules_class['timewindow']['big'] += 1
        else:
            violation_rules_class['timewindow']['treshold'] += 1
            treshold_related_case.add(_GT['src_info']['hash'])

    elif vl == 'rule4_fee' and 'generate_key_info_fail' not in _GT['violate_rule']:
        fee_rate, src_symbol, src_addr, dst_symbol, dst_addr = ( 
            _GT['pair_info']['fee_rate'],
            _GT['src_info']['token_symbol'], 
            _GT['src_info']['token_addr'], 
            _GT['dst_info']['token_symbol'], 
            _GT['dst_info']['token_addr']
        )
        violation_rules_statis['fee_rate'].append((fee_rate, src_symbol, src_addr, dst_symbol, dst_addr))
        symbol_equal = (src_symbol == dst_symbol) or ((src_symbol, dst_symbol) in (("USD", "eth"), ("eth", "USD")) )
        if symbol_equal and fee_rate > 0.01: 
            violation_rules_class['fee_rate']['treshold'] += 1
            treshold_related_case.add(_GT['src_info']['hash'])
        elif symbol_equal and fee_rate < 0: 
            violation_rules_class['fee_rate']['neg'] += 1
        elif (isinstance(src_addr, str) and src_addr.startswith('0x42000')) or (isinstance(dst_addr, str) and  dst_addr.startswith('0x42000')):
            violation_rules_class['fee_rate']['op'] += 1
        elif not symbol_equal: 
            violation_rules_class['fee_rate']['symbol_mismatch'] += 1
        else:
            violation_rules_class['fee_rate']['other'] += 1
    elif vl == 'rule3_token_name':
        src_symbol, dst_symbol = (_GT['src_info']['token_symbol'], _GT['dst_info']['token_symbol'])
        violation_rules_statis['token_symbol'][
            f"{src_symbol}->{dst_symbol}"
        ] += 1
        src_type = DOC.get_token_symbol_type(src_symbol)
        dst_type = DOC.get_token_symbol_type(dst_symbol)
        violation_rules_class['token_symbol'][f"{src_type}->{dst_type}"] +=1

def _get_info_from_two_tx_hash(tx1_hash:str, tx2_hash:str, tx_hash_to_key_info:T.Dict)->T.Dict:
    # help analyze result only
    tx1_info = ed(tx_hash_to_key_info.get(tx1_hash, {}))
    tx2_info = ed(tx_hash_to_key_info.get(tx2_hash, {}))
    info = dict()

    # violate rules
    info['violate_rule'] = []
    if len(tx1_info) and len(tx2_info):
        rules = get_rules()
        for rule in rules:
            if not rule(tx1_info, tx2_info):
                info['violate_rule'].append(rule.__name__)
    else:
        info['violate_rule'].append('generate_key_info_fail')
    info['src_info'] = tx1_info
    info['dst_info'] = tx2_info

    # pair info
    pair_info = {}
    try:
        ts1 = TL.save_value_int(tx1_info.get('timestamp', 0))
    except :
        ts1 = 0
    try:
        ts2 = TL.save_value_int(tx2_info.get('timestamp', 0))
    except :
        ts2 = 0
    amount1 = tx1_info.get('amount', 0)
    amount2 = tx2_info.get('amount', 0)
    
    if not (isinstance(amount1, (int, float))):
        amount1 = 0
    if not isinstance(amount2,(int,float)):
        amount2 = 0

    if ts1 is None or ts2 is None:
        pair_info['time_window'] = float('nan')
    else:
        pair_info['time_window'] = ts2 - ts1
    if amount1 == 0:
        pair_info['fee_rate'] = float('nan')
        pair_info['fee_amount'] = amount1 - amount2
    else:
        pair_info['fee_rate'] = (amount1 - amount2) / amount1
        pair_info['fee_amount'] = amount1 - amount2
    info['pair_info'] = pair_info

    return info


def generate_FP_plus():
    print('---generate FP_plus begin---')

    # 看下FP中，me vs gt
    # 关注 time window，fee，fee异常时的amount
    FP = dict()
    with open(C.paths().FP, 'r') as f:
        FP = json.load(f)

    tx_hash_to_key_info = _generate_tx_hash_to_key_info_dict()
    FP_rules = dd(lambda :0)
    FP_plus = dict()
    for src_chain in FP:
        FP_plus[src_chain] = {}
        for dst_chain in FP[src_chain]:
            FP_plus[src_chain][dst_chain] = []
            for vs in FP[src_chain][dst_chain]:
                me_src_tx = vs['me'][0]
                me_dst_tx = vs['me'][1]
                gt_dst_tx = vs['gt'][1] # gt_dst_tx can be ''
                
                vs['me_info'] = _get_info_from_two_tx_hash(me_src_tx, me_dst_tx, tx_hash_to_key_info)
                vs['gt_info'] = _get_info_from_two_tx_hash(me_src_tx, gt_dst_tx, tx_hash_to_key_info) if len(gt_dst_tx) else {}
                if not len(vs['gt_info']['violate_rule']): 
                    FP_rules['empty'] += 1
                else:
                    for vl in vs['gt_info']['violate_rule']: 
                        FP_rules[vl] += 1
                FP_plus[src_chain][dst_chain].append(vs)
    
    print(TL.defaultdict_to_dict(FP_rules))
    with open(C.paths().FP_plus, 'w') as f:
        json.dump(FP_plus, f, indent=2)
                
    print('---generate FP_plus end---')


def generate_FN_plus():
    print('---generate FN_plus begin---')

    # 看下FN中，违背了哪些rule
    FN = dict()
    with open(C.paths().FN, 'r') as f:
        FN = json.load(f)

    all_timestamp = dd(dict)
    for chain in FN:
        data = TL.load_all_data_by_chain(chain)
        for trx in data: 
            tx_hash = trx['hash'].lower()
            all_timestamp[chain][tx_hash] = TL.save_value_int(trx.get('timestamp') or trx.get('timeStamp'))

    tx_hash_to_key_info = _generate_tx_hash_to_key_info_dict()
    FN_rules = {
        'cnt': dd(lambda :0),
        'statis':{
            'timewindow': [],
            'fee_rate': [],
            'token_symbol':dd(lambda:0),
            'empty':0
        },
        'class': {
            'timewindow': dd(lambda:0),
            'fee_rate': dd(lambda:0),
            'token_symbol':dd(lambda:0)
        }
    }
    violation_rules_cnt = FN_rules['cnt']
    violation_rules_statis = FN_rules['statis']
    violation_rules_class = FN_rules['class']
    FN_plus = dict()
    for src_chain in FN:
        FN_plus[src_chain] = {}
        for dst_chain in FN[src_chain]:
            FN_plus[src_chain][dst_chain] = []
            for gt in FN[src_chain][dst_chain]:
                gt_src_tx = gt[0].lower()
                gt_dst_tx = gt[1].lower()
                
                gt_plus = {}
                gt_plus['src_tx'] = gt_src_tx
                gt_plus['dst_tx'] = gt_dst_tx
                gt_plus['info'] = _get_info_from_two_tx_hash(gt_src_tx, gt_dst_tx, tx_hash_to_key_info)
                if not len(gt_plus['info']['violate_rule']):
                    violation_rules_statis['empty'] += 1
                for vl in gt_plus['info']['violate_rule']: 
                    classify_violation_rules(vl, 
                            gt_plus['info'], violation_rules_cnt, violation_rules_statis,
                            violation_rules_class, all_timestamp[src_chain][gt_src_tx], all_timestamp[dst_chain][gt_dst_tx])

                FN_plus[src_chain][dst_chain].append(gt_plus)
    
    print(f"treshold_related_case:{len(treshold_related_case)}")
    with open(C.paths().FN_plus, 'w') as f:
        json.dump(FN_plus, f, indent=2)
    with open(C.paths().FN_rules , 'w') as f:
        json.dump(FN_rules, f, indent=2)
    print('---generate FN_plus end---')


def generate_TP_plus():
    print('---generate TP_plus begin---')

    # 看下TP
    TP = dict()
    with open(C.paths().TP, 'r') as f:
        TP = json.load(f)
   
    tx_hash_to_key_info = _generate_tx_hash_to_key_info_dict()

    TP_plus = dict()
    for src_chain in TP:
        TP_plus[src_chain] = {}
        for dst_chain in TP[src_chain]:
            TP_plus[src_chain][dst_chain] = []
            for gt in TP[src_chain][dst_chain]:
                gt_src_tx = gt[0]
                gt_dst_tx = gt[1]
                
                gt_plus = {}
                gt_plus['src_tx'] = gt_src_tx
                gt_plus['dst_tx'] = gt_dst_tx
                gt_plus['info'] = _get_info_from_two_tx_hash(gt_src_tx, gt_dst_tx, tx_hash_to_key_info)
                TP_plus[src_chain][dst_chain].append(gt_plus)
    
    with open(C.paths().TP_plus, 'w') as f:
        json.dump(TP_plus, f, indent=2)
                
    print('---generate TP_plus end---')


def analyze_FN_plus_token():
    print('---analyze FN_plus token begin---')
    # 分析FN_plus中，本该匹配但没匹配的token
    FN_plus = dict()
    paired_token:T.List[T.Dict] = []
    
    with open(C.paths().FN_plus, 'r') as f:
        FN_plus = json.load(f)

    for src_chain in FN_plus:
        for dst_chain in FN_plus[src_chain]:
            for one_gt in FN_plus[src_chain][dst_chain]:
                src_info = one_gt['info']['src_info']
                dst_info = one_gt['info']['dst_info']
                src_token_info = _generate_token_info_from_key_info(src_info)
                dst_token_info = _generate_token_info_from_key_info(dst_info)
                fee_rate = one_gt['info']['pair_info']['fee_rate']
                fee_amount = one_gt['info']['pair_info']['fee_amount']
                # filter token with different name, now useless
                #if src_token_info['token_name'] == dst_token_info['token_name']: 
                #    continue
                one_pair = {
                    'token1': src_token_info, 
                    'token2': dst_token_info, 
                    'count': 1,
                    'fee_rate': [fee_rate],
                    'fee_amount': [fee_amount]
                }
                
                is_exist = False  # 去重，不需要传递性
                for other_pair in paired_token:
                    t1, t2 = (one_pair['token1'], one_pair['token2'])
                    t3, t4 = (other_pair['token1'], other_pair['token2'])
                    if (t1, t2) == (t3, t4) or (t2, t1) == (t3, t4):
                        is_exist = True
                        other_pair['count'] += 1
                        other_pair['fee_rate'].append(fee_rate)
                        other_pair['fee_rate'].sort()
                        other_pair['fee_amount'].append(fee_amount)
                        other_pair['fee_amount'].sort()
                        break
                if not is_exist:
                    paired_token.append(one_pair)

    with open(C.paths().analysis_FN_token, 'w') as f:
        json.dump(paired_token, f, indent=2)

    _write_paired_token_to_excel(paired_token, C.paths().analysis_FN_token_excel)
    
    print('---analyze FN_plus token end---')
